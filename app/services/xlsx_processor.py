import os
import json
import re
import hashlib
from typing import Optional
from app.services.aes_gcm import encrypt_with_metadata, generate_key as generate_key_aes
from openpyxl import load_workbook
from app.services.pii_main import extract_all_pii

def mask_xlsx_sensitive_text(xlsx_path: str, key_path: Optional[str] = None, enabled_pii_categories=None):
    if key_path is None:
        key_path = xlsx_path.replace(".xlsx", ".key")

    if os.path.exists(key_path):
        with open(key_path, "rb") as f:
            key = f.read()
    else:
        key = generate_key_aes()
        with open(key_path, "wb") as f:
            f.write(key)

    wb = load_workbook(xlsx_path)

    full_text = ""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    full_text += cell.value + " "

    # Tag format: [ENC:{Label}_{8-char-hash}]
    # Examples: [ENC:NAMES_abcd1234], [ENC:Email_wxyz5678], [ENC:Phone_defg9012]
    all_pii_list = extract_all_pii(full_text, enabled_pii_categories)

    # Filter by enabled categories
    non_selectable_categories = ['IC', 'Email', 'DOB', 'Bank Account', 'Passport', 'Phone', 'Credit Card', 'Address', 'Vehicle Registration']
    filtered_pii_list = []
    for label, value in all_pii_list:
        if label in enabled_pii_categories or label in non_selectable_categories:
            filtered_pii_list.append((label, value))
            print(f"[FILTER] Masking XLSX: {label} = {value}")
        else:
            print(f"[FILTER] Skipping XLSX (not enabled): {label} = {value}")

    all_pii_list = filtered_pii_list
    print(f"print all_pii_list: '{all_pii_list}'")
    
    unique_pii = {}
    masked_pii = []

    for label, value in all_pii_list:
        if value not in unique_pii:
            try:
                encrypted = encrypt_with_metadata(value, key)
                encrypted_str = json.dumps(encrypted)
                value_hash = hashlib.md5(value.encode()).hexdigest()[:8]
                unique_tag = f"[ENC:{label}_{value_hash}]"

                unique_pii[value] = {"encrypted": encrypted_str, "tag": unique_tag}
                masked_pii.append({
                    "original": value,
                    "encrypted": encrypted_str,
                    "label": label,
                    "masked": unique_tag
                })
            except Exception as e:
                print(f"[ERROR] Failed to encrypt '{value}': {e}")

    sorted_pii_items = sorted(unique_pii.items(), key=lambda x: len(x[0]), reverse=True)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    masked_text = cell.value

                    for pii_value, pii_info in sorted_pii_items:
                        # Use word boundary regex to replace only complete words
                        escaped_pii = re.escape(pii_value)
                        pattern = r'\b' + escaped_pii + r'\b'
                        masked_text = re.sub(pattern, pii_info["tag"], masked_text)

                    cell.value = masked_text

    output_path = xlsx_path.replace(".xlsx", ".masked.xlsx")
    wb.save(output_path)
    
    json_output_path = xlsx_path.replace(".xlsx", ".masked.json")
    with open(json_output_path, "w", encoding="utf-8") as f:
        json.dump(masked_pii, f, indent=2)

    return {
        "status": "success",
        "masked_file": output_path,
        "json_output": json_output_path,
        "key_file": key_path
    }

def run_xlsx_processing(xlsx_path: str, enabled_pii_categories=None):
    try:
        key_path = xlsx_path.replace(".xlsx", ".key")
        result = mask_xlsx_sensitive_text(xlsx_path, key_path, enabled_pii_categories)
        return {
            "status": result.get("status", "error"),
            "masked_xlsx": result.get("masked_file"),
            "json_output": result.get("json_output"),
            "key_file": result.get("key_file")
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }
